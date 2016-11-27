import sys
import openpyxl as xl


def interpolate(T, low_properties, high_properties):

	# If both are the same, no need to interpolate
	if low_properties['Temp'] == high_properties['Temp']:
		return low_properties

	# Determine proportionality factor
	T_low = low_properties['Temp']
	T_high = high_properties['Temp']
	factor = (T-T_low)/(T_high-T_low)

	# Interpolate each property
	retval = {}
	for key, value in low_properties.items():
		if value is None:
			retval[key] = None
		else:
			retval[key] = factor*(high_properties[key]-low_properties[key]) + low_properties[key]

	return retval


def load_sheets(fluid):

	# Load the excel sheet corresponding to the fluid
	if fluid == 'Water':
		wb = xl.load_workbook('data/Water.xlsx')
	elif fluid == 'R134a':
		wb = xl.load_workbook('data/R134a.xlsx', read_only = True)
	elif fluid == 'Ammonia':
		wb = xl.load_workbook('data/Ammonia.xlsx', read_only = True)
	elif fluid == 'Propane':
		wb = xl.load_workbook('data/Propane.xlsx', read_only = True)

	# Separate constants from properties
	constants = wb['Constants']
	properties = wb['Properties']

	return {'Constants': constants, 'Properties': properties}



def find_indexes(T, properties):

	# Iterate through column A from index 5 to the point where the temperature is lower than the wanted one
	i = 4
	current_temp = properties['A5'].value;

	if current_temp > T:
		raise Exception('Error: temperature out of range. (Too low)')

	while current_temp <= T:
		i = i+1
		key = 'A%i' % i
		current_temp = properties[key].value

		if current_temp is not None:
			# Check if the column already exists
			if T == current_temp:
				return [i, i]
		else:
			raise Exception('Temperature out of range. (Too high)')

	return [i-1, i]


def get_low_high(indexes, properties):

	# Range of cells in which the properties to interpolate are
	rg = 'A%i:U%i' % (indexes[0], indexes[1])

	# Get all the cells
	data = properties[rg]

	cells  = []
	for row in properties.iter_rows(rg):
		for cell in row:
			cells.append(cell.value)

	low_properties = {
		'Temp': cells[0],
		'p': cells[1],
		'rho_L' : cells[2],
		'rho_G' : cells[3],
		'u_L' : cells[4],
		'u_LG': cells [5],
		'h_L' :cells[6],
		'h_LG' : cells[7],
		's_L' : cells[8],
		's_LG' : cells[9],
		'cp_L' : cells[10],
		'cp_G' : cells[11],
		'k_L' : cells[12],
		'k_G' : cells[13],
		'mu_L' : cells[14],
		'mu_G' : cells[15],
		'sigma' : cells[16],
		'nu_L' : cells[14]/cells[2],
		'nu_G' : cells[15]/cells[3],
		'alpha_L' : cells[12]*cells[2]/cells[10],
		'alpha_G' : cells[13]*cells[3]/cells[11]
	}

	if indexes[0] == indexes[1]:
		high_properties = low_properties
	else:
		high_properties = {
		'Temp': cells[21],
		'p': cells[22],
		'rho_L' : cells[23],
		'rho_G' : cells[24],
		'u_L' : cells[25],
		'u_LG': cells [26],
		'h_L' :cells[27],
		'h_LG' : cells[28],
		's_L' : cells[29],
		's_LG' : cells[30],
		'cp_L' : cells[31],
		'cp_G' : cells[32],
		'k_L' : cells[33],
		'k_G' : cells[34],
		'mu_L' : cells[35],
		'mu_G' : cells[36],
		'sigma' : cells[37],
		'nu_L' : cells[35]/cells[23],
		'nu_G' : cells[36]/cells[24],
		'alpha_L' : cells[33]*cells[23]/cells[31],
		'alpha_G' : cells[34]*cells[24]/cells[32]
		}

	return (low_properties, high_properties)


def get_properties(T, fluid):

	message = 'Getting properties at %0.2fK for %s' % (T, fluid)
	#print(message)

	ws = load_sheets(fluid)
	indexes = find_indexes(T, ws['Properties'])
	rows = get_low_high(indexes, ws['Properties'])

	# Interpolated properties
	properties = interpolate(T, rows[0], rows[1])

	# Add constants
	values = ws['Constants']

	properties['C_sf'] = values['B1'].value
	properties['fluid'] = values['B2'].value
	properties['M'] = values['B3'].value
	properties['p_crit'] = values['B4'].value
	properties['alpha_0'] = values['B5'].value

	return properties
